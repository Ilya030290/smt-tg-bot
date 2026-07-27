import os
import math
import tempfile
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import logging

logger = logging.getLogger(__name__)

def calculate_offsets(board_width, board_height, x_count, y_count, pitch_x, pitch_y, rotated_blocks):
    offsets = []
    total = x_count * y_count

    def format_num(n):
        if isinstance(n, float) and n.is_integer():
            return str(int(n))
        return str(n)

    if total == 2 and len(rotated_blocks) == 1:
        for n in range(1, total + 1):
            if n in rotated_blocks:
                offsets.append(f"{format_num(board_width)};{format_num(board_height)};180")
            else:
                offsets.append("0;0;0")
        return offsets

    for n in range(1, total + 1):
        row = (n - 1) // x_count
        col = (n - 1) % x_count
        if n in rotated_blocks:
            x_off = board_width - ((x_count - 1 - col) * pitch_x)
            y_off = board_height - ((y_count - 1 - row) * pitch_y)
            r = 180
        else:
            x_off = col * pitch_x
            y_off = row * pitch_y
            r = 0
        offsets.append(f"{format_num(x_off)};{format_num(y_off)};{r}")
    return offsets

def apply_delta_to_dataframe(df, delta_x, delta_y):
    df['Center-X(mm)'] = df['Center-X(mm)'] + delta_x
    df['Center-Y(mm)'] = df['Center-Y(mm)'] + delta_y
    return df

def generate_pnp_from_xlsm(xlsm_path, output_dir, project_name, pcb_side):
    wb = openpyxl.load_workbook(xlsm_path, data_only=True)
    ws_proj = wb['Project Data']
    ws_pnp = wb['PNPwizard']

    board_dim_raw = ws_proj['B3'].value
    board_dim = str(board_dim_raw).replace(' ', '')
    parts = board_dim.split(';')
    board_width = float(parts[0]) if parts else 0

    # Исправленный формат: два знака после запятой
    offset_x = -(board_width - 5)
    offset_y = -5
    board_offset = f"{offset_x:.2f};{offset_y:.2f}"

    block_offsets = []
    row = 18
    while True:
        val = ws_proj.cell(row=row, column=2).value
        if val is None:
            break
        block_offsets.append(str(val).replace(' ', ''))
        row += 1

    side = pcb_side.upper()
    if side == 'BOT':
        fid1 = ws_proj['B6'].value if ws_proj['B6'].value is not None else '0;0'
        fid2 = ws_proj['B7'].value if ws_proj['B7'].value is not None else '0;0'
    else:
        fid1 = ws_proj['B8'].value if ws_proj['B8'].value is not None else '0;0'
        fid2 = ws_proj['B9'].value if ws_proj['B9'].value is not None else '0;0'
    fiducials = [str(fid1).replace(' ', ''), str(fid2).replace(' ', '')]

    mount_data = []
    for row in ws_pnp.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        refdes = str(row[0]) if row[0] is not None else ''
        part = str(row[1]) if row[1] is not None else ''
        x = row[2] if row[2] is not None else 0
        y = row[3] if row[3] is not None else 0
        rot = row[4] if row[4] is not None else 0
        x_str = f"{x:.6f}".rstrip('0').rstrip('.') if isinstance(x, (int, float)) else str(x)
        y_str = f"{y:.6f}".rstrip('0').rstrip('.') if isinstance(y, (int, float)) else str(y)
        rot_str = f"{rot:.6f}".rstrip('0').rstrip('.') if isinstance(rot, (int, float)) else str(rot)
        mount_data.append(f"{refdes};{part};{x_str};{y_str};{rot_str}")

    lines = []
    lines.append("# Board dimensions (X, Y, Thickness):")
    lines.append(board_dim)
    lines.append("")
    lines.append("# Board offset (X, Y):")
    lines.append(board_offset)
    lines.append("")
    lines.append("# Block offset (X, Y, Rotation):")
    lines.extend(block_offsets)
    lines.append("")
    lines.append("# Fiducials (X, Y):")
    lines.extend(fiducials)
    lines.append("")
    lines.append("# Mount data (RefDes, Partnumber, X, Y, Rotation):")
    lines.extend(mount_data)
    lines.append("")

    pnp_filename = f"{project_name}_{pcb_side}.pnp"
    pnp_path = os.path.join(output_dir, pnp_filename)
    with open(pnp_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))

    return pnp_path
