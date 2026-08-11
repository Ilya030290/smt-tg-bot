import pandas as pd
import re
import openpyxl
from openpyxl.styles import Font, PatternFill
import tempfile
import os


def parse_pnp_to_dataframe(pnp_path):
    
    with open(pnp_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    mount_data = []
    in_mount = False

    for line in lines:
        if line.startswith('# Mount data'):
            in_mount = True
            continue

        if in_mount and line.strip() and not line.startswith('#'):
            parts = line.strip().split(';')

            if len(parts) >= 5:
                refdes = parts[0]
                partnumber = parts[1]
                x = parts[2]
                y = parts[3]
                rot = parts[4]

                mount_data.append([
                    refdes,
                    partnumber,
                    x,
                    y,
                    rot
                ])

    df = pd.DataFrame(
        mount_data,
        columns=[
            'RefDes',
            'Partnumber',
            'X',
            'Y',
            'Rotation'
        ]
    )

    return df


def parse_csv_to_dataframe(csv_path):

    header_line_index = None

    with open(
        csv_path,
        'r',
        encoding='utf-8-sig',
        errors='replace'
    ) as f:

        lines = f.readlines()

    for index, line in enumerate(lines):

        line_clean = line.strip()

        if (
            line_clean.startswith('No.,')
            and 'Pattern Name' in line_clean
            and 'Part Name' in line_clean
            and ',X,' in line_clean
            and ',Y,' in line_clean
            and ',R,' in line_clean
        ):
            header_line_index = index
            break

    if header_line_index is None:
        raise ValueError(
            "Не удалось найти заголовок таблицы в CSV. "
            "Ожидались колонки 'Pattern Name', 'Part Name', 'X', 'Y' и 'R'."
        )

    df = pd.read_csv(
        csv_path,
        skiprows=header_line_index,
        sep=',',
        dtype=str,
        encoding='utf-8-sig',
        engine='python'
    )

    
    df.columns = [
        str(column).strip()
        for column in df.columns
    ]

    required_columns = [
        'Pattern Name',
        'Part Name',
        'X',
        'Y',
        'R'
    ]

    missing_columns = [
        column
        for column in required_columns
        if column not in df.columns
    ]

    if missing_columns:
        raise ValueError(
            "В CSV отсутствуют необходимые колонки: "
            + ", ".join(missing_columns)
        )

    result_df = pd.DataFrame()

    result_df['RefDes'] = (
        df['Pattern Name']
        .fillna('')
        .astype(str)
        .str.strip()
    )

    result_df['Partnumber'] = (
        df['Part Name']
        .fillna('')
        .astype(str)
        .str.strip()
    )

    result_df['X'] = (
        df['X']
        .fillna('')
        .astype(str)
        .str.strip()
    )

    result_df['Y'] = (
        df['Y']
        .fillna('')
        .astype(str)
        .str.strip()
    )

    result_df['Rotation'] = (
        df['R']
        .fillna('')
        .astype(str)
        .str.strip()
    )

    result_df = result_df[
        (result_df['RefDes'] != '') |
        (result_df['Partnumber'] != '')
    ].copy()


    result_df.reset_index(drop=True, inplace=True)

    return result_df


def normalize_value(value):
    if not isinstance(value, str):
        value = str(value)

    return (
        value
        .replace(" ", "")
        .replace(".", ",")
        .upper()
    )


def validate_pnp_with_bom(pnp_df, bom_df):

    bom_df = bom_df.copy()
    pnp_df = pnp_df.copy()

    bom_df['Positions'] = (
        bom_df['Positions']
        .astype(str)
        .str.strip()
    )

    bom_df['Article name'] = (
        bom_df['Article name']
        .astype(str)
        .str.strip()
    )

    pnp_df['RefDes'] = (
        pnp_df['RefDes']
        .astype(str)
        .str.strip()
    )

    pnp_df['Partnumber'] = (
        pnp_df['Partnumber']
        .astype(str)
        .str.strip()
    )

    bom_dict = {}

    for _, row in bom_df.iterrows():

        positions = re.split(
            r',\s*',
            row['Positions']
        )

        for pos in positions:

            pos = pos.strip()

            if pos:
                bom_dict[pos] = row['Article name']


    result_rows = []

    for _, pnp_row in pnp_df.iterrows():

        refdes = pnp_row['RefDes']
        partnumber = pnp_row['Partnumber']

        if refdes in bom_dict:

            article = bom_dict[refdes]

            match = (
                normalize_value(partnumber)
                ==
                normalize_value(article)
            )

        else:

            article = ''
            match = False

        result_rows.append({
            'RefDes': refdes,
            'Partnumber': partnumber,
            'Positions': refdes if refdes in bom_dict else '',
            'Article name': article,
            'Соответствие': 'Да' if match else 'Нет'
        })

    return pd.DataFrame(result_rows)


def generate_validation_report(result_df):

    output_dir = tempfile.gettempdir()

    output_path = os.path.join(
        output_dir,
        'validation_report.xlsx'
    )

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'Report'

    headers = [
        'RefDes',
        'Partnumber',
        'Positions',
        'Article name',
        'Соответствие'
    ]

    ws.append(headers)

    for _, row in result_df.iterrows():

        ws.append([
            row['RefDes'],
            row['Partnumber'],
            row['Positions'],
            row['Article name'],
            row['Соответствие']
        ])

    green_fill = PatternFill(
        start_color='00B050',
        end_color='00B050',
        fill_type='solid'
    )

    red_fill = PatternFill(
        start_color='FF0000',
        end_color='FF0000',
        fill_type='solid'
    )

    bold_font = Font(bold=True)

    for row_idx in range(
        2,
        len(result_df) + 2
    ):

        cell = ws.cell(
            row=row_idx,
            column=5
        )

        if cell.value == 'Да':
            cell.fill = green_fill

        elif cell.value == 'Нет':
            cell.fill = red_fill

    for col in range(1, 6):

        ws.cell(
            row=1,
            column=col
        ).font = bold_font


    for col in ws.columns:

        max_length = 0

        column = col[0].column_letter

        for cell in col:

            try:

                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))

            except Exception:
                pass

        adjusted_width = min(
            max_length + 2,
            30
        )

        ws.column_dimensions[
            column
        ].width = adjusted_width

    wb.save(output_path)
    wb.close()

    return output_path
