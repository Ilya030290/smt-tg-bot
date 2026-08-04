import os
import tempfile
import re
import pandas as pd
from openpyxl.styles import Font

def apply_formatting_to_results(sheet, data_rows):
    """Форматирует лист Results (цвета и итоговая строка)."""
    headers = {}
    for col_idx, cell in enumerate(sheet[1], start=1):
        if cell.value == "Article name":
            headers["article"] = col_idx
        elif cell.value == "Qty":
            headers["qty"] = col_idx
        elif cell.value == "Qty Positions":
            headers["qty_pos"] = col_idx
        elif cell.value == "Correspondence of Qty":
            headers["corresp"] = col_idx

    if not headers:
        return

    green_font = Font(color="00B050")
    red_font = Font(color="FF0000")
    for row_idx in range(2, data_rows + 2):
        cell = sheet.cell(row=row_idx, column=headers["corresp"])
        if cell.value == "Соответствует":
            cell.font = green_font
        elif cell.value == "Не соответствует":
            cell.font = red_font

    mismatches = []
    for row_idx in range(2, data_rows + 2):
        corr_cell = sheet.cell(row=row_idx, column=headers["corresp"])
        if corr_cell.value == "Не соответствует":
            article = sheet.cell(row=row_idx, column=headers["article"]).value
            qty = sheet.cell(row=row_idx, column=headers["qty"]).value
            qty_pos = sheet.cell(row=row_idx, column=headers["qty_pos"]).value
            mismatches.append((article, qty, qty_pos))

    start_row = data_rows + 3
    title_cell = sheet.cell(row=start_row, column=1)
    title_cell.value = "Результат соответствия по количеству:"
    title_cell.font = Font(bold=True)

    if mismatches:
        row_offset = start_row + 1
        sheet.cell(row=row_offset, column=1).value = "Article name"
        sheet.cell(row=row_offset, column=2).value = "Qty"
        sheet.cell(row=row_offset, column=3).value = "Qty Positions"
        for r, (article, qty, qty_pos) in enumerate(mismatches, start=row_offset + 1):
            sheet.cell(row=r, column=1).value = article
            sheet.cell(row=r, column=2).value = qty
            sheet.cell(row=r, column=3).value = qty_pos
            for c in range(1, 4):
                sheet.cell(row=r, column=c).font = red_font
    else:
        msg_cell = sheet.cell(row=start_row + 1, column=1)
        msg_cell.value = "Несоответствий не выявлено"
        msg_cell.font = Font(color="00B050", bold=True)

def transform_pnp(input_path, original_filename):
    """
    Преобразует Excel-файл в PNP-формат.
    input_path – путь к входному файлу (временный)
    original_filename – оригинальное имя файла
    Возвращает путь к созданному выходному файлу.
    """
    df = pd.read_excel(input_path, header=None,
                       names=["Article name", "Qty", "Positions"])
    df = df[~df["Positions"].astype(str).str.lower().eq("positions")]
    df = df[~df["Qty"].astype(str).str.lower().eq("qty")]
    df["Qty"] = pd.to_numeric(df["Qty"], errors="coerce")

    mask = (
        df["Positions"].isna() |
        df["Positions"].astype(str).str.strip().eq("") |
        df["Positions"].astype(str).str.lower().eq("nan")
    ) & (
        df["Article name"].astype(str).str.contains("PCB", case=False, na=False)
    )
    df_filtered = df[~mask].copy()

    transformed_data = []
    for _, row in df_filtered.iterrows():
        article = row["Article name"]
        positions = str(row["Positions"]).split(",")
        for pos in positions:
            transformed_data.append({"Positions": pos.strip(), "Article name": article})
    transformed_df = pd.DataFrame(transformed_data)

    transformed_df = transformed_df[~transformed_df["Positions"].str.lower().eq("positions")]
    mask2 = (
        transformed_df["Positions"].isna() |
        transformed_df["Positions"].eq("") |
        transformed_df["Positions"].eq("nan")
    ) & (
        transformed_df["Article name"].astype(str).str.contains("PCB", case=False, na=False)
    )
    transformed_df = transformed_df[~mask2]

    results_data = []
    for _, row in df_filtered.iterrows():
        article = row["Article name"]
        qty = row["Qty"]
        pos_str = str(row["Positions"]) if pd.notna(row["Positions"]) else ""
        if pos_str.strip() == "":
            qty_positions = 0
        else:
            pos_list = [p.strip() for p in pos_str.split(",") if p.strip() != ""]
            qty_positions = len(pos_list)
        correspondence = "Соответствует" if qty == qty_positions else "Не соответствует"
        results_data.append({
            "Article name": article,
            "Qty": qty,
            "Qty Positions": qty_positions,
            "Correspondence of Qty": correspondence
        })
    results_df = pd.DataFrame(results_data)

    base_name = os.path.splitext(original_filename)[0]
    safe_base_name = re.sub(r'[\\/*?:"<>|]', '_', base_name)
    output_dir = tempfile.gettempdir()
    output_path = os.path.join(output_dir, f"{safe_base_name}_PNP.xlsx")

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        transformed_df.to_excel(writer, sheet_name="PNP", index=False)
        results_df.to_excel(writer, sheet_name="Results", index=False)
        workbook = writer.book
        sheet = workbook["Results"]
        apply_formatting_to_results(sheet, len(results_df))

    return output_path
