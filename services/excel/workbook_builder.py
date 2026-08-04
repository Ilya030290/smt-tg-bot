import os
import math
import tempfile
import pandas as pd
import xlwings as xw
import openpyxl
import logging
from services.excel.block_offsets import calculate_offsets

logger = logging.getLogger(__name__)

def apply_delta_to_dataframe(df, delta_x, delta_y):
    df['Center-X(mm)'] = df['Center-X(mm)'] + delta_x
    df['Center-Y(mm)'] = df['Center-Y(mm)'] + delta_y
    return df

def generate_pnp_from_xlsm(xlsm_path, output_dir, project_name, pcb_side):
    """Генерирует .pnp файл из готового .xlsm."""
    wb = openpyxl.load_workbook(xlsm_path, data_only=True)
    ws_proj = wb['Project Data']
    ws_pnp = wb['PNPwizard']

    board_dim_raw = ws_proj['B3'].value
    board_dim = str(board_dim_raw).replace(' ', '')
    parts = board_dim.split(';')
    board_width = float(parts[0]) if parts else 0

    offset_x = -(board_width - 5)
    offset_y = -5
    board_offset = f"{offset_x:.2f};{offset_y:.2f}"

    block_offsets = []
    row = 18
    while True:
        val = ws_proj.cell(row=row, column=2).value
        if val is None:
            break
        block_offsets.append(str(val).replace(' ', ''))
        row += 1

    side = pcb_side.upper()
    if side == 'BOT':
        fid1 = ws_proj['B6'].value if ws_proj['B6'].value is not None else '0;0'
        fid2 = ws_proj['B7'].value if ws_proj['B7'].value is not None else '0;0'
    else:
        fid1 = ws_proj['B8'].value if ws_proj['B8'].value is not None else '0;0'
        fid2 = ws_proj['B9'].value if ws_proj['B9'].value is not None else '0;0'
    fiducials = [str(fid1).replace(' ', ''), str(fid2).replace(' ', '')]

    mount_data = []
    for row in ws_pnp.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        refdes = str(row[0]) if row[0] is not None else ''
        part = str(row[1]) if row[1] is not None else ''
        x = row[2] if row[2] is not None else 0
        y = row[3] if row[3] is not None else 0
        rot = row[4] if row[4] is not None else 0
        x_str = f"{x:.6f}".rstrip('0').rstrip('.') if isinstance(x, (int, float)) else str(x)
        y_str = f"{y:.6f}".rstrip('0').rstrip('.') if isinstance(y, (int, float)) else str(y)
        rot_str = f"{rot:.6f}".rstrip('0').rstrip('.') if isinstance(rot, (int, float)) else str(rot)
        mount_data.append(f"{refdes};{part};{x_str};{y_str};{rot_str}")

    lines = []
    lines.append("# Board dimensions (X, Y, Thickness):")
    lines.append(board_dim)
    lines.append("")
    lines.append("# Board offset (X, Y):")
    lines.append(board_offset)
    lines.append("")
    lines.append("# Block offset (X, Y, Rotation):")
    lines.extend(block_offsets)
    lines.append("")
    lines.append("# Fiducials (X, Y):")
    lines.extend(fiducials)
    lines.append("")
    lines.append("# Mount data (RefDes, Partnumber, X, Y, Rotation):")
    lines.extend(mount_data)
    lines.append("")

    pnp_filename = f"{project_name}_{pcb_side}.pnp"
    pnp_path = os.path.join(output_dir, pnp_filename)
    with open(pnp_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))

    return pnp_path

def build_program_files(project_data, df, template_path):
    """
    Основная функция генерации .xlsm и .pnp.
    project_data – экземпляр ProjectData
    df – pandas DataFrame с данными
    template_path – путь к шаблону .xlsm
    Возвращает (xlsm_path, pnp_path)
    """
    # Сортировка
    df = df.sort_values(by='Positions').reset_index(drop=True)

    # Добавление SN-LABEL, если нужно
    if project_data.need_sn_label and project_data.sn_label_coords:
        x, y, angle = map(float, project_data.sn_label_coords.split(";"))
        sn_row = {
            "Positions": "SN-LABEL",
            "Article name": "SN-LABEL",
            "Center-X(mm)": x,
            "Center-Y(mm)": y,
            "Rotation": angle,
        }
        df = pd.concat([df, pd.DataFrame([sn_row])], ignore_index=True)

    # Создание временного файла .xlsm
    output_dir = tempfile.gettempdir()
    filename = f"{project_data.project_name}_{project_data.pcb_side}.xlsm"
    output_path = os.path.join(output_dir, filename)

    with xw.App(visible=False, add_book=False) as app:
        wb = app.books.open(template_path)

        ws_proj = wb.sheets['Project Data']
        ws_proj.range('B1').value = project_data.project_name
        ws_proj.range('B2').value = project_data.pcb_side
        ws_proj.range('B3').value = project_data.board_dimensions
        ws_proj.range('B4').value = project_data.multiplication
        if project_data.rotated_blocks:
            ws_proj.range('B5').value = '0;0'
        else:
            ws_proj.range('B5').value = f"{project_data.pitch_x};{project_data.pitch_y}"

        side = project_data.pcb_side.upper()
        if side == 'BOT':
            ws_proj.range('B6').value = project_data.fiducial_bot1
            ws_proj.range('B7').value = project_data.fiducial_bot2
        else:
            ws_proj.range('B8').value = project_data.fiducial_top1
            ws_proj.range('B9').value = project_data.fiducial_top2

        board_x, board_y, _ = map(float, project_data.board_dimensions.split(';'))
        mult_x, mult_y = map(int, project_data.multiplication.split(';'))
        pitch_x = project_data.pitch_x
        pitch_y = project_data.pitch_y
        rotated = project_data.rotated_blocks
        offsets = calculate_offsets(board_x, board_y, mult_x, mult_y, pitch_x, pitch_y, rotated)
        for i, offset in enumerate(offsets):
            ws_proj.range(f'B{18 + i}').value = offset

        ws_pnp = wb.sheets['PNPwizard']
        for idx, row in df.iterrows():
            row_num = idx + 2
            ws_pnp.range(f'A{row_num}').value = row['Positions']
            ws_pnp.range(f'B{row_num}').value = row['Article name']
            ws_pnp.range(f'C{row_num}').value = row['Center-X(mm)']
            ws_pnp.range(f'D{row_num}').value = row['Center-Y(mm)']
            ws_pnp.range(f'E{row_num}').value = row['Rotation']

        wb.save(output_path)
        wb.close()

    # Генерация .pnp
    try:
        pnp_path = generate_pnp_from_xlsm(output_path, output_dir, project_data.project_name, project_data.pcb_side)
    except Exception as e:
        logger.error(f"Ошибка при генерации PNP: {e}")
        pnp_path = None

    return output_path, pnp_path
