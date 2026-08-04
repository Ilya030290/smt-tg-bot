import os
import tempfile
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime

def clean_article_name(value):
    if pd.isna(value):
        return value
    str_value = str(value)
    str_value = str_value.replace(',', '.')
    str_value = str_value.replace(' ', '_')
    return str_value

def process_rotation(position, rotation_value):
    if pd.isna(position):
        return 0
    pos_str = str(position).strip()
    if pos_str and pos_str[0] in ['R', 'C', 'L']:
        try:
            rot_val = float(rotation_value) if not pd.isna(rotation_value) else 0
            rot_val = round(rot_val)
            if rot_val in [0, 180]:
                return 0
            elif rot_val in [90, 270]:
                return 90
            else:
                return 0
        except:
            return 0
    else:
        return 0

def merge_tables(file1_path, file2_path, original_filename1=None):
    # Чтение
    df1 = pd.read_excel(file1_path)
    df2 = pd.read_excel(file2_path)

    # Проверка столбцов
    if 'Positions' not in df1.columns or 'Article name' not in df1.columns:
        raise ValueError("Таблица1 должна содержать столбцы 'Positions' и 'Article name'")
    required = ['Designator', 'Layer', 'Center-X(mm)', 'Center-Y(mm)', 'Rotation']
    for col in required:
        if col not in df2.columns:
            raise ValueError(f"Таблица2 должна содержать столбец '{col}'")

    # Очистка Article name
    df1['Article name'] = df1['Article name'].apply(clean_article_name)

    # Сортировка
    df1_sorted = df1.sort_values(by='Positions').reset_index(drop=True)
    df2_sorted = df2.sort_values(by='Designator').reset_index(drop=True)

    # DNP: компоненты из df2, которых нет в df1
    positions_set = set(df1_sorted['Positions'])
    designator_set = set(df2_sorted['Designator'])
    non_matching_designators = list(designator_set - positions_set)

    df2_working = df2_sorted.copy()
    dnp_bot_list = []
    dnp_top_list = []

    for designator in non_matching_designators:
        row = df2_working[df2_working['Designator'] == designator]
        if not row.empty:
            layer = str(row.iloc[0]['Layer'])
            if 'bot' in layer.lower():
                dnp_bot_list.append(designator)
            else:
                dnp_top_list.append(designator)
            df2_working = df2_working[df2_working['Designator'] != designator]

    # --- НОВОЕ: находим позиции из df1, которых нет в df2 ---
    df1_sorted = df1_sorted.sort_values(by='Positions').reset_index(drop=True)
    df2_working = df2_working.sort_values(by='Designator').reset_index(drop=True)

    # Выполняем слияние
    result_df = pd.merge(df1_sorted, df2_working, left_on='Positions', right_on='Designator', how='left')

    # Находим строки, у которых Designator (из df2) отсутствует
    unmatched_mask = result_df['Designator'].isna()
    unmatched_positions = result_df.loc[unmatched_mask, 'Positions'].tolist()

    # Удаляем эти строки из результирующего DataFrame
    result_df = result_df[~unmatched_mask].copy()

    # Добавляем unmatched_positions в DNP_BOT с пометкой
    for pos in unmatched_positions:
        dnp_bot_list.append(f"{pos} (в Odin есть, но не найден в PNP заказчика)")

    # Удаляем лишний столбец Designator
    result_df.drop('Designator', axis=1, inplace=True)

    # Rotation
    def fix_rotation(pos, rot):
        if pd.isna(pos):
            return 0
        if str(pos).strip() and str(pos).strip()[0] in 'RCL':
            try:
                val = float(rot) if not pd.isna(rot) else 0
                val = round(val)
                if val in (0, 180):
                    return 0
                elif val in (90, 270):
                    return 90
            except:
                pass
        return 0

    result_df['Rotation'] = result_df.apply(lambda row: fix_rotation(row['Positions'], row['Rotation']), axis=1)

    # Порядок столбцов
    cols = ['Positions', 'Article name', 'Center-X(mm)', 'Center-Y(mm)', 'Rotation', 'Layer']
    result_df = result_df[cols]

    # Разделение по слоям
    result_df['Layer'] = result_df['Layer'].astype(str)
    bottom_df = result_df[result_df['Layer'].str.lower().str.contains('bottom', na=False)]
    top_df = result_df[~result_df['Layer'].str.lower().str.contains('bottom', na=False)]

    # DNP DataFrame
    max_len = max(len(dnp_bot_list), len(dnp_top_list))
    dnp_data = []
    for i in range(max_len):
        bot = dnp_bot_list[i] if i < len(dnp_bot_list) else ''
        top = dnp_top_list[i] if i < len(dnp_top_list) else ''
        dnp_data.append({'DNP_BOT': bot, 'DNP_TOP': top})
    dnp_df = pd.DataFrame(dnp_data)

    # ---- Сохранение во временную папку ----
    output_dir = tempfile.gettempdir()
    if original_filename1:
        base_name = os.path.splitext(os.path.basename(original_filename1))[0]
        if "_from" in base_name:
            base_name = base_name.split("_from")[0]
    else:
        base_name = "merged_result"
    safe_name = re.sub(r'[\\/*?:"<>|]', '_', base_name)
    output_path = os.path.join(output_dir, f"{safe_name}_result.xlsx")

    if os.path.exists(output_path):
        os.remove(output_path)

    with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
        bottom_df.to_excel(writer, sheet_name='BottomLayer', index=False)
        top_df.to_excel(writer, sheet_name='TopLayer', index=False)
        dnp_df.to_excel(writer, sheet_name='DNP_List', index=False)

    if not os.path.exists(output_path) or os.path.getsize(output_path) == 0:
        raise RuntimeError("Файл не создан или пуст")

    # Форматирование DNP_List (красным) через openpyxl
    try:
        wb = load_workbook(output_path)
        ws = wb['DNP_List']
        red = Font(color="FF0000")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                if cell.value:
                    cell.font = red
        info = f"Всего DNP_BOT: {len(dnp_bot_list)} | DNP_TOP: {len(dnp_top_list)}"
        ws['E1'] = info
        ws['E1'].font = Font(bold=True)
        wb.save(output_path)
        wb.close()
    except Exception as e:
        print(f"Предупреждение: не удалось применить форматирование: {e}")

    return output_path
